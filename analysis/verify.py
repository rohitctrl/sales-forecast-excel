"""
Structural verification of the generated .xlsx.

openpyxl will happily re-read its own output even when Excel would show a
"repair" prompt, so a reload alone proves nothing. These checks are stricter:

  1. the zip container passes CRC verification
  2. every XML part is well-formed
  3. the parts Excel requires are present and declared in [Content_Types].xml
  4. every internal relationship Target resolves to a part that exists
  5. openpyxl can reload the file and the expected sheets and formulas are there

LibreOffice is not installed on the build machine, so no headless-convert check
is performed. That limitation is stated in the README rather than papered over.
"""

from __future__ import annotations

import os
import posixpath
import zipfile
import xml.etree.ElementTree as ET

import openpyxl

REQUIRED_PARTS = [
    "[Content_Types].xml",
    "_rels/.rels",
    "xl/workbook.xml",
    "xl/_rels/workbook.xml.rels",
    "xl/styles.xml",
]
EXPECTED_SHEETS = ["Notes", "Data", "Lookups", "Dashboard", "Forecast"]

HEADER_ROW = 5
FIRST_ROW = 6
N_MONTHS = 24
N_HOLDOUT = 3
TRAIN_LAST_ROW = FIRST_ROW + (N_MONTHS - N_HOLDOUT) - 1
DESIGN_FIRST_COL = 21   # column U: t + 11 month dummies
DESIGN_N_COLS = 12
ACTUAL_COL = 3


def _design_matrix_check(fc, expected_pred) -> dict:
    """
    Refit OLS from the numbers actually written into the workbook's design
    matrix and compare against the forecast run_analysis.py published. This
    catches a mis-laid-out sheet; it cannot prove how Excel evaluates TREND.
    """
    import numpy as np

    rows = range(FIRST_ROW, FIRST_ROW + N_MONTHS)
    X = np.array([[fc.cell(row=r, column=DESIGN_FIRST_COL + j).value
                   for j in range(DESIGN_N_COLS)] for r in rows], dtype=float)
    y = np.array([fc.cell(row=r, column=ACTUAL_COL).value for r in rows], dtype=float)
    X = np.column_stack([np.ones(len(X)), X])
    ntr = N_MONTHS - N_HOLDOUT
    beta, *_ = np.linalg.lstsq(X[:ntr], y[:ntr], rcond=None)
    pred = X[ntr:] @ beta
    diff = float(np.max(np.abs(pred - np.asarray(expected_pred, float))))
    return {"ok": diff < 0.5,
            "detail": f"max abs difference GBP {diff:.6f} between the sheet's own "
                      "design matrix refitted and the published forecast"}


def verify(path: str, expected_pred=None) -> dict:
    res = {"path": os.path.abspath(path), "bytes": os.path.getsize(path),
           "checks": {}, "ok": True}

    def check(name, passed, detail=""):
        res["checks"][name] = {"pass": bool(passed), "detail": detail}
        if not passed:
            res["ok"] = False

    with zipfile.ZipFile(path) as zf:
        bad = zf.testzip()
        check("zip_crc_valid", bad is None, f"first bad member: {bad}" if bad else "all members OK")

        names = set(zf.namelist())
        check("required_parts_present",
              all(p in names for p in REQUIRED_PARTS),
              "missing: " + ", ".join(p for p in REQUIRED_PARTS if p not in names))

        malformed = []
        for n in sorted(names):
            if n.endswith(".xml") or n.endswith(".rels"):
                try:
                    ET.fromstring(zf.read(n))
                except ET.ParseError as e:
                    malformed.append(f"{n}: {e}")
        check("all_xml_well_formed", not malformed, "; ".join(malformed[:5]))

        dangling = []
        ns = "{http://schemas.openxmlformats.org/package/2006/relationships}"
        for n in sorted(names):
            if not n.endswith(".rels"):
                continue
            base = posixpath.dirname(posixpath.dirname(n))
            root = ET.fromstring(zf.read(n))
            for rel in root.findall(f"{ns}Relationship"):
                if rel.get("TargetMode") == "External":
                    continue
                tgt = rel.get("Target", "")
                if tgt.startswith("/"):
                    resolved = tgt.lstrip("/")
                else:
                    resolved = posixpath.normpath(posixpath.join(base, tgt))
                if resolved not in names:
                    dangling.append(f"{n} -> {tgt}")
        check("no_dangling_relationships", not dangling, "; ".join(dangling[:5]))

        ct = zf.read("[Content_Types].xml").decode("utf-8")
        sheet_parts = sorted(p for p in names if p.startswith("xl/worksheets/sheet"))
        declared = all(posixpath.basename(p).replace(".xml", "") in ct or f"/{p}" in ct
                       for p in sheet_parts)
        check("worksheets_declared", declared and len(sheet_parts) == len(EXPECTED_SHEETS),
              f"{len(sheet_parts)} worksheet parts")

    wb = openpyxl.load_workbook(path)
    check("reload_sheetnames", wb.sheetnames == EXPECTED_SHEETS, str(wb.sheetnames))
    check("full_recalc_on_load", bool(wb.calculation.fullCalcOnLoad),
          "workbook forces recalculation when opened")

    dash = wb["Dashboard"]
    formulas = [c.value for row in dash.iter_rows() for c in row
                if isinstance(c.value, str) and c.value.startswith("=")]
    check("dashboard_has_sumifs", sum("SUMIFS(" in f for f in formulas) >= 100,
          f"{sum('SUMIFS(' in f for f in formulas)} SUMIFS formulas")
    check("dashboard_has_vlookup", sum("VLOOKUP(" in f for f in formulas) >= 12,
          f"{sum('VLOOKUP(' in f for f in formulas)} VLOOKUP formulas")
    check("dashboard_no_xlookup", not any("XLOOKUP" in f for f in formulas),
          "XLOOKUP needs an _xlfn prefix through openpyxl and would render #NAME?")
    check("dashboard_conditional_formatting",
          len(list(dash.conditional_formatting)) >= 3,
          f"{len(list(dash.conditional_formatting))} conditional formatting ranges")
    check("dashboard_charts", len(dash._charts) >= 2, f"{len(dash._charts)} charts")

    fc = wb["Forecast"]
    ff = [c.value for row in fc.iter_rows() for c in row
          if isinstance(c.value, str) and c.value.startswith("=")]
    check("forecast_fits_with_linest", sum("LINEST(" in f for f in ff) >= 13,
          f"{sum('LINEST(' in f for f in ff)} LINEST formulas")
    check("forecast_predicts_with_trend", sum("TREND(" in f for f in ff) >= 24,
          f"{sum('TREND(' in f for f in ff)} TREND formulas "
          "(prediction does not depend on LINEST coefficient ordering)")
    check("forecast_charts", len(fc._charts) >= 1, f"{len(fc._charts)} charts")
    check("forecast_holdout_not_in_training",
          all(f"$C${TRAIN_LAST_ROW}" in f for f in ff if f.startswith("=TREND(")),
          f"every TREND known_y stops at row {TRAIN_LAST_ROW}, "
          "so the held-out months never enter the fit")

    dm = _design_matrix_check(fc, expected_pred)
    check("workbook_design_matrix_reproduces_python", dm["ok"], dm["detail"])

    data = wb["Data"]
    check("data_is_values_not_formulas",
          not any(isinstance(c.value, str) and c.value.startswith("=")
                  for row in data.iter_rows(min_row=2) for c in row),
          "source facts are literal values, as they should be")
    check("data_table_defined", len(data.tables) == 1, str(list(data.tables)))
    return res
