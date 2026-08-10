"""
Build the deliverable .xlsx with openpyxl.

What is genuinely live in this file (recalculated by Excel/LibreOffice, not pasted):
  * Dashboard  - every KPI, matrix cell and product row is a SUMIFS / SUM / VLOOKUP
                 formula reading the Data sheet. A region dropdown drives the KPI row.
  * Forecast   - the regression is fitted INSIDE the workbook with LINEST over the
                 21 training rows; fitted values, forecasts and MAPE / RMSE / R^2 are
                 formulas. Nothing is hardcoded from Python.
  * Formatting - colour scale on the region x month matrix, data bars on product revenue.
  * Charts     - native openpyxl.chart objects bound to those formula cells.

What is NOT here, honestly: openpyxl cannot emit a native PivotTable (no pivotCache /
pivotTable part writer). The pivot logic is implemented with SUMIFS against a tidy
fact table instead. See the Notes sheet.
"""

from __future__ import annotations

import os
from datetime import datetime, timezone

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

FIXED_TIMESTAMP = datetime(2025, 2, 1, 0, 0, 0, tzinfo=timezone.utc).replace(tzinfo=None)

INK = "FF141414"
BLUE = "FF1F2CD1"
PLATE = "FFECEAE2"
GREY = "FFB9B6AC"

H1 = Font(name="Calibri", size=14, bold=True, color=INK)
H2 = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
BOLD = Font(name="Calibri", size=10, bold=True, color=INK)
BODY = Font(name="Calibri", size=10, color=INK)
SMALL = Font(name="Calibri", size=9, color="FF5A5A5A")
HDR_FILL = PatternFill("solid", fgColor=BLUE)
PLATE_FILL = PatternFill("solid", fgColor=PLATE)
THIN = Side(style="thin", color=GREY)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

GBP = '#,##0'
GBP2 = '#,##0.00'
PCT2 = '0.00%'
NUM = '#,##0'


def _title(ws, row, text, sub=None):
    ws.cell(row=row, column=1, value=text).font = H1
    if sub:
        ws.cell(row=row + 1, column=1, value=sub).font = SMALL
    return row + (2 if sub else 1)


def _header_row(ws, row, values, start_col=1):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=start_col + i, value=v)
        c.font = H2
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
    ws.row_dimensions[row].height = 26


def build(path, fact, months, regions, top_codes, product_names,
          series, holdout, rolling, provenance, ledger_steps, claim_notes):
    wb = openpyxl.Workbook()
    # openpyxl stamps docProps/core.xml with datetime.now() by default, which makes
    # the output bytes differ on every run. Pin it so the file is reproducible and
    # its sha256 can be published.
    wb.properties.creator = "run_analysis.py (openpyxl)"
    wb.properties.lastModifiedBy = "run_analysis.py (openpyxl)"
    wb.properties.created = FIXED_TIMESTAMP
    wb.properties.modified = FIXED_TIMESTAMP
    wb.properties.title = "Online Retail II - sales dashboard and forecast"
    wb.calculation.fullCalcOnLoad = True

    notes = wb.active
    notes.title = "Notes"
    data_ws = wb.create_sheet("Data")
    look_ws = wb.create_sheet("Lookups")
    dash_ws = wb.create_sheet("Dashboard")
    fc_ws = wb.create_sheet("Forecast")

    n_fact = _build_data(data_ws, fact)
    _build_lookups(look_ws, months, regions, top_codes, product_names, n_fact)
    _build_dashboard(dash_ws, months, regions, top_codes, n_fact)
    _build_forecast(fc_ws, series, holdout, rolling)
    _build_notes(notes, provenance, ledger_steps, claim_notes, n_fact, len(months))

    wb.active = wb.index(dash_ws)
    wb.save(path)
    _normalise(path)
    return path


def _normalise(path):
    """
    Make the saved file byte-reproducible.

    openpyxl overwrites docProps/core.xml's <dcterms:modified> with the wall clock
    at save time, and zip member timestamps drift, so two runs over identical data
    produce different bytes. Rewrite the container with the modified stamp pinned
    and every member dated 1980-01-01, so the workbook's sha256 is a real integrity
    check rather than a timestamp.
    """
    import re
    import shutil
    import tempfile
    import zipfile

    stamp = FIXED_TIMESTAMP.strftime("%Y-%m-%dT%H:%M:%SZ")
    with zipfile.ZipFile(path) as src:
        members = [(i, src.read(i.filename)) for i in src.infolist()]

    fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=os.path.dirname(path) or ".")
    os.close(fd)
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as out:
        for info, payload in members:
            if info.filename == "docProps/core.xml":
                payload = re.sub(
                    rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                    rb"\g<1>" + stamp.encode() + rb"\g<2>", payload)
            zi = zipfile.ZipInfo(info.filename, date_time=(1980, 1, 1, 0, 0, 0))
            zi.compress_type = zipfile.ZIP_DEFLATED
            zi.external_attr = info.external_attr
            out.writestr(zi, payload)
    shutil.move(tmp, path)
    os.chmod(path, 0o644)


# --------------------------------------------------------------------- Data

DATA_COLS = ["Month", "Region", "ProductCode", "ProductName",
             "Revenue", "Units", "Orders", "Year", "MonthNo"]


def _build_data(ws, fact):
    _header_row(ws, 1, DATA_COLS)
    for i, rec in enumerate(fact.itertuples(index=False), start=2):
        y, m = str(rec.month).split("-")
        vals = [str(rec.month), rec.region, rec.product, rec.product_name,
                round(float(rec.revenue), 2), int(rec.units), int(rec.orders),
                int(y), int(m)]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = BODY
        ws.cell(row=i, column=5).number_format = GBP2
        ws.cell(row=i, column=6).number_format = NUM
        ws.cell(row=i, column=7).number_format = NUM
    n = len(fact) + 1
    tbl = Table(displayName="FactSales", ref=f"A1:I{n}")
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight9", showRowStripes=True, showColumnStripes=False)
    ws.add_table(tbl)
    for col, w in zip("ABCDEFGHI", [10, 16, 13, 38, 14, 10, 10, 8, 9]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    return n


# ------------------------------------------------------------------ Lookups

def _build_lookups(ws, months, regions, top_codes, product_names, n_fact):
    r = _title(ws, 1, "Lookup tables",
               "VLOOKUP targets and the lists that drive the Dashboard dropdown.")
    r += 1
    _header_row(ws, r, ["ProductCode", "ProductName", "TotalRevenue"])
    first = r + 1
    for i, code in enumerate(top_codes + ["ALL OTHER PRODUCTS"]):
        rr = first + i
        ws.cell(row=rr, column=1, value=code).font = BODY
        ws.cell(row=rr, column=2,
                value=product_names.get(code, "All other products")).font = BODY
        c = ws.cell(row=rr, column=3,
                    value=f"=SUMIFS(Data!$E$2:$E${n_fact},Data!$C$2:$C${n_fact},$A{rr})")
        c.number_format = GBP
        c.font = BODY
    ws["E3"] = "Region"
    ws["E3"].font = BOLD
    ws["E4"] = "All regions"
    ws["E4"].font = BODY
    for i, reg in enumerate(regions):
        ws.cell(row=5 + i, column=5, value=reg).font = BODY
    ws["G3"] = "Month"
    ws["G3"].font = BOLD
    for i, m in enumerate(months):
        ws.cell(row=4 + i, column=7, value=m).font = BODY
    for col, w in zip("ABCDEFG", [16, 38, 15, 3, 18, 3, 12]):
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------- Dashboard

def _build_dashboard(ws, months, regions, top_codes, n_fact):
    D = f"Data!$A$2:$A${n_fact}"      # month
    Rg = f"Data!$B$2:$B${n_fact}"     # region
    Pc = f"Data!$C$2:$C${n_fact}"     # product code
    Rv = f"Data!$E$2:$E${n_fact}"     # revenue
    Un = f"Data!$F$2:$F${n_fact}"     # units
    Or = f"Data!$G$2:$G${n_fact}"     # orders

    _title(ws, 1, "Online Retail II - sales dashboard",
           "Every figure below is a live formula over the Data sheet. "
           "Change the region filter and the whole KPI row recalculates.")

    ws["A4"] = "Region filter"
    ws["A4"].font = BOLD
    ws["B4"] = "All regions"
    ws["B4"].font = Font(name="Calibri", size=10, bold=True, color=BLUE)
    ws["B4"].fill = PLATE_FILL
    ws["B4"].border = BOX
    dv = DataValidation(type="list", formula1="Lookups!$E$4:$E$8", allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(ws["B4"])

    kpis = [
        ("Total revenue (GBP)", f'=IF($B$4="All regions",SUM({Rv}),SUMIFS({Rv},{Rg},$B$4))', GBP),
        ("Units sold", f'=IF($B$4="All regions",SUM({Un}),SUMIFS({Un},{Rg},$B$4))', NUM),
        ("Order lines", f'=IF($B$4="All regions",SUM({Or}),SUMIFS({Or},{Rg},$B$4))', NUM),
        ("Peak month revenue", f'=MAX(B10:B33)', GBP),
        ("Average month revenue", f'=AVERAGE(B10:B33)', GBP),
        ("Peak / trough ratio", f'=MAX(B10:B33)/MIN(B10:B33)', '0.00"x"'),
    ]
    for i, (label, formula, fmt) in enumerate(kpis):
        col = 4 + i * 2
        lc = ws.cell(row=4, column=col, value=label)
        lc.font = SMALL
        lc.alignment = Alignment(horizontal="left", wrap_text=True)
        vc = ws.cell(row=5, column=col, value=formula)
        vc.font = Font(name="Calibri", size=13, bold=True, color=BLUE)
        vc.number_format = fmt
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)

    # ---- monthly x region matrix (the "pivot", built from SUMIFS)
    ws["A8"] = "Revenue by month and region"
    ws["A8"].font = BOLD
    _header_row(ws, 9, ["Month", "Total (filtered)"] + regions + ["MoM growth"])
    first_m = 10
    for i, m in enumerate(months):
        r = first_m + i
        ws.cell(row=r, column=1, value=m).font = BODY
        tot = ws.cell(
            row=r, column=2,
            value=(f'=IF($B$4="All regions",SUMIFS({Rv},{D},$A{r}),'
                   f'SUMIFS({Rv},{D},$A{r},{Rg},$B$4))'))
        tot.number_format = GBP
        tot.font = BOLD
        for j, reg in enumerate(regions):
            c = ws.cell(row=r, column=3 + j,
                        value=f'=SUMIFS({Rv},{D},$A{r},{Rg},"{reg}")')
            c.number_format = GBP
            c.font = BODY
        g = ws.cell(row=r, column=3 + len(regions),
                    value=("" if i == 0 else f'=IFERROR(B{r}/B{r-1}-1,"")'))
        g.number_format = PCT2
        g.font = BODY
    last_m = first_m + len(months) - 1

    matrix = f"C{first_m}:{get_column_letter(2 + len(regions))}{last_m}"
    ws.conditional_formatting.add(matrix, ColorScaleRule(
        start_type="min", start_color="FFFFFFFF",
        mid_type="percentile", mid_value=50, mid_color="FFAEB2EE",
        end_type="max", end_color=BLUE))
    ws.conditional_formatting.add(f"B{first_m}:B{last_m}", DataBarRule(
        start_type="min", end_type="max", color="1F2CD1", showValue=True))
    ws.conditional_formatting.add(
        f"{get_column_letter(3 + len(regions))}{first_m}:"
        f"{get_column_letter(3 + len(regions))}{last_m}",
        ColorScaleRule(start_type="min", start_color="FFF2B8B8",
                       mid_type="num", mid_value=0, mid_color="FFFFFFFF",
                       end_type="max", end_color="FFB8E0C2"))

    # ---- top products block (SUMIFS + VLOOKUP)
    pr0 = last_m + 3
    ws.cell(row=pr0 - 1, column=1, value="Top 12 products by revenue "
            "(of 4,893 distinct stock codes)").font = BOLD
    _header_row(ws, pr0, ["ProductCode", "Product name (VLOOKUP)", "Revenue",
                          "Units", "Share of total", "Revenue per unit"])
    for i, code in enumerate(top_codes):
        r = pr0 + 1 + i
        ws.cell(row=r, column=1, value=code).font = BODY
        ws.cell(row=r, column=2,
                value=f'=VLOOKUP($A{r},Lookups!$A:$B,2,FALSE)').font = BODY
        c = ws.cell(row=r, column=3, value=f'=SUMIFS({Rv},{Pc},$A{r})')
        c.number_format = GBP
        c.font = BODY
        u = ws.cell(row=r, column=4, value=f'=SUMIFS({Un},{Pc},$A{r})')
        u.number_format = NUM
        u.font = BODY
        s = ws.cell(row=r, column=5, value=f'=C{r}/SUM({Rv})')
        s.number_format = PCT2
        s.font = BODY
        p = ws.cell(row=r, column=6, value=f'=IFERROR(C{r}/D{r},"")')
        p.number_format = GBP2
        p.font = BODY
    pr_last = pr0 + len(top_codes)
    ws.conditional_formatting.add(f"C{pr0+1}:C{pr_last}", DataBarRule(
        start_type="num", start_value=0, end_type="max", color="1F2CD1",
        showValue=True))

    # ---- native charts bound to the formula cells
    line = LineChart()
    line.title = "Monthly revenue by region"
    line.style = 2
    line.y_axis.title = "Revenue (GBP)"
    line.x_axis.title = "Month"
    data = Reference(ws, min_col=3, max_col=2 + len(regions),
                     min_row=9, max_row=last_m)
    cats = Reference(ws, min_col=1, min_row=first_m, max_row=last_m)
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    line.height, line.width = 8.5, 22
    ws.add_chart(line, f"A{pr_last + 3}")

    bar = BarChart()
    bar.type = "bar"
    bar.title = "Top 12 products by revenue"
    bar.y_axis.title = "Revenue (GBP)"
    bdata = Reference(ws, min_col=3, min_row=pr0, max_row=pr_last)
    bcats = Reference(ws, min_col=2, min_row=pr0 + 1, max_row=pr_last)
    bar.add_data(bdata, titles_from_data=True)
    bar.set_categories(bcats)
    bar.height, bar.width = 9.5, 22
    ws.add_chart(bar, f"A{pr_last + 21}")

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 38
    for col in "CDEFGHIJKLMNO":
        ws.column_dimensions[col].width = 16
    ws.freeze_panes = "A10"


# ----------------------------------------------------------------- Forecast

def _build_forecast(ws, series, holdout, rolling):
    """
    The regression is fitted by Excel's LINEST inside the workbook, over the
    training rows only. Python's numbers are never written into the model.
    """
    n = len(series)
    n_test = len(holdout["test_months"])
    n_train = n - n_test

    r = _title(ws, 1, "Forecast - trend + month-of-year regression",
               "LINEST fits the model on the first "
               f"{n_train} months only. The last {n_test} months are held out; "
               "the workbook never sees them during fitting.")

    hdr = ["Month", "t", "Actual", "Set"] + \
          [f"m{m:02d}" for m in range(2, 13)] + \
          ["Regression", "Seasonal naive", "APE reg", "APE snaive"]
    HR = 5
    _header_row(ws, HR, hdr)
    first = HR + 1
    last = first + n - 1
    train_last = first + n_train - 1
    test_first = train_last + 1

    # Column layout: A Month, B t, C Actual, D Set, E..O dummies m02..m12,
    # P Regression, Q Seasonal naive, R APE reg, S APE snaive.
    # LINEST needs ONE contiguous known_x block, and columns C/D sit between t and
    # the dummies, so a mirrored contiguous design matrix is kept off to the right.
    DM = len(hdr) + 2                       # column U: t + 11 dummies = 12 columns
    dm_first = get_column_letter(DM)
    dm_last = get_column_letter(DM + 11)
    CB = DM + 14                            # coefficient block, clear of the matrix
    cbL = get_column_letter(CB)
    cbV = get_column_letter(CB + 1)

    y_rng = f"$C${first}:$C${train_last}"
    x_rng = f"${dm_first}${first}:${dm_last}${train_last}"

    ws.cell(row=HR, column=DM, value="t (design)").font = H2
    ws.cell(row=HR, column=DM).fill = HDR_FILL
    for k in range(11):
        c = ws.cell(row=HR, column=DM + 1 + k, value=f"x{k+2:02d}")
        c.font = H2
        c.fill = HDR_FILL

    for i, (period, value) in enumerate(series.items()):
        row = first + i
        is_train = i < n_train
        ws.cell(row=row, column=1, value=str(period)).font = BODY
        ws.cell(row=row, column=2, value=i).font = BODY
        a = ws.cell(row=row, column=3, value=round(float(value), 2))
        a.number_format = GBP
        a.font = BODY
        s = ws.cell(row=row, column=4, value="train" if is_train else "HOLDOUT")
        s.font = BOLD if not is_train else BODY
        for k, m in enumerate(range(2, 13)):
            ws.cell(row=row, column=5 + k,
                    value=1 if period.month == m else 0).font = BODY
        # mirrored contiguous design matrix
        ws.cell(row=row, column=DM, value=i)
        for k, m in enumerate(range(2, 13)):
            ws.cell(row=row, column=DM + 1 + k, value=1 if period.month == m else 0)

        # Fitted value / forecast. TREND fits the same OLS as LINEST but returns
        # the prediction directly, so the cell does not depend on Excel's
        # reverse coefficient ordering - one less thing that can silently break.
        # known_y and known_x are the TRAINING rows only; new_x is this row.
        p = ws.cell(row=row, column=16,
                    value=(f"=TREND({y_rng},{x_rng},"
                           f"${dm_first}{row}:${dm_last}{row})"))
        p.number_format = GBP
        p.font = Font(name="Calibri", size=10, bold=not is_train, color=BLUE)

        if i >= 12:
            sn = ws.cell(row=row, column=17, value=f"=C{row-12}")
        else:
            sn = ws.cell(row=row, column=17, value="")
        sn.number_format = GBP
        sn.font = BODY

        e1 = ws.cell(row=row, column=18, value=f"=ABS((C{row}-P{row})/C{row})")
        e1.number_format = PCT2
        e1.font = BODY
        e2 = ws.cell(row=row, column=19,
                     value=(f"=IF(Q{row}=\"\",\"\",ABS((C{row}-Q{row})/C{row}))"))
        e2.number_format = PCT2
        e2.font = BODY

    # ---- LINEST coefficient block (for reading the model, not for predicting)
    ws.cell(row=HR - 1, column=CB, value="Coefficients (Excel LINEST, "
            f"training rows {first}-{train_last}). LINEST returns coefficients in "
            "reverse column order with the intercept last; these cells unpick that. "
            "The forecast column uses TREND and does not depend on this block."
            ).font = BOLD
    names = ["Intercept", "Trend (per month)"] + \
            [f"Month {m:02d} vs Jan" for m in range(2, 13)]
    # LINEST returns coefficients right-to-left across known_x, intercept last.
    n_x = 12
    for i, nm in enumerate(names):
        row = _coef_row(i)
        ws.cell(row=row, column=CB, value=nm).font = BODY
        if i == 0:
            idx = n_x + 1                       # intercept is the final entry
        else:
            idx = n_x - (i - 1)                 # x1 -> n_x, x2 -> n_x-1, ...
        c = ws.cell(row=row, column=CB + 1,
                    value=f"=INDEX(LINEST({y_rng},{x_rng},TRUE,FALSE),1,{idx})")
        c.number_format = GBP2
        c.font = BODY

    # ---- metric block
    mrow = _coef_row(len(names)) + 2
    ws.cell(row=mrow - 1, column=CB,
            value=f"Holdout metrics ({n_test} months, rows {test_first}-{last})").font = BOLD
    tf, tl = test_first, last
    metrics = [
        ("MAPE - regression", f"=AVERAGE(R{tf}:R{tl})", PCT2),
        ("MAPE - seasonal naive", f"=AVERAGE(S{tf}:S{tl})", PCT2),
        ("RMSE - regression", f"=SQRT(SUMXMY2(C{tf}:C{tl},P{tf}:P{tl})/{n_test})", GBP2),
        ("RMSE - seasonal naive", f"=SQRT(SUMXMY2(C{tf}:C{tl},Q{tf}:Q{tl})/{n_test})", GBP2),
        ("R2 - regression",
         f"=1-SUMXMY2(C{tf}:C{tl},P{tf}:P{tl})/DEVSQ(C{tf}:C{tl})", "0.000"),
        ("R2 - seasonal naive",
         f"=1-SUMXMY2(C{tf}:C{tl},Q{tf}:Q{tl})/DEVSQ(C{tf}:C{tl})", "0.000"),
        ("In-sample R2 (LINEST)",
         f"=INDEX(LINEST({y_rng},{x_rng},TRUE,TRUE),3,1)", "0.000"),
    ]
    for i, (nm, f, fmt) in enumerate(metrics):
        row = mrow + i
        ws.cell(row=row, column=CB, value=nm).font = BODY
        c = ws.cell(row=row, column=CB + 1, value=f)
        c.number_format = fmt
        c.font = Font(name="Calibri", size=10, bold=True, color=BLUE)

    vrow = mrow + len(metrics) + 2
    ws.cell(row=vrow, column=CB, value="Cross-check - values computed in Python "
            "by run_analysis.py (these two blocks must agree)").font = BOLD
    checks = [
        ("MAPE regression (%)", round(holdout["models"]["linreg_dummies"]["mape_pct"], 4)),
        ("MAPE seasonal naive (%)", round(holdout["models"]["seasonal_naive"]["mape_pct"], 4)),
        ("RMSE regression", round(holdout["models"]["linreg_dummies"]["rmse"], 4)),
        ("RMSE seasonal naive", round(holdout["models"]["seasonal_naive"]["rmse"], 4)),
        ("R2 regression", round(holdout["models"]["linreg_dummies"]["r2"], 4)),
        ("R2 seasonal naive", round(holdout["models"]["seasonal_naive"]["r2"], 4)),
        ("Rolling-origin MAPE regression (%)",
         round(rolling["models"]["linreg_dummies"]["mape_pct"], 4)),
        ("Rolling-origin MAPE seasonal naive (%)",
         round(rolling["models"]["seasonal_naive"]["mape_pct"], 4)),
    ]
    coefs = holdout["raw"]["linreg_dummies_coef"]
    checks += [("Python coefficient: intercept", round(coefs[0], 4)),
               ("Python coefficient: trend per month", round(coefs[1], 4))]
    checks += [(f"Python coefficient: month {m:02d} vs Jan", round(coefs[2 + k], 4))
               for k, m in enumerate(range(2, 13))]
    for i, (nm, v) in enumerate(checks):
        ws.cell(row=vrow + 1 + i, column=CB, value=nm).font = SMALL
        ws.cell(row=vrow + 1 + i, column=CB + 1, value=v).font = SMALL

    # ---- chart: actual vs fitted vs forecast
    ch = LineChart()
    ch.title = "Actual vs fitted vs held-out forecast"
    ch.y_axis.title = "Revenue (GBP)"
    ch.x_axis.title = "Month"
    ch.add_data(Reference(ws, min_col=3, min_row=HR, max_row=last), titles_from_data=True)
    ch.add_data(Reference(ws, min_col=16, min_row=HR, max_row=last), titles_from_data=True)
    ch.add_data(Reference(ws, min_col=17, min_row=HR, max_row=last), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=first, max_row=last))
    ch.height, ch.width = 9, 24
    ws.add_chart(ch, f"A{last + 3}")

    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["D"].width = 10
    for col in ["P", "Q", "R", "S"]:
        ws.column_dimensions[col].width = 15
    ws.column_dimensions[cbL].width = 34
    ws.column_dimensions[cbV].width = 18
    ws.freeze_panes = "A6"


def _coef_row(i: int) -> int:
    return 6 + i


# -------------------------------------------------------------------- Notes

def _build_notes(ws, provenance, ledger_steps, claim_notes, n_fact, n_months):
    r = 1
    ws.cell(row=r, column=1, value="Read this first").font = H1
    r += 2
    blocks = [
        ("What this workbook is",
         ["A sales dashboard and a monthly revenue forecast built on the UCI "
          "Online Retail II transaction file.",
          "The Dashboard and Forecast sheets contain live formulas, not pasted "
          "values. Open it and press F9 - the numbers rebuild from the Data sheet.",
          "Authoritative numbers are produced by analysis/run_analysis.py in the "
          "repository. This workbook reproduces them independently with LINEST."]),
        ("Honest note on PivotTables",
         ["This file contains NO PivotTables. openpyxl cannot write a pivotCache "
          "or pivotTable part, so claiming one would be false.",
          "The pivot logic - revenue sliced by month, region and product - is "
          "implemented with SUMIFS against a tidy fact table on the Data sheet, "
          "plus a data-validation dropdown that re-filters the KPI row.",
          "For the reader's purpose this is equivalent: it produces the same "
          "cross-tab, it recalculates live when the source data changes, and "
          "unlike a PivotTable it does not need a manual refresh."]),
        ("Data grain",
         [f"The Data sheet holds {n_fact - 1:,} aggregated rows at "
          "month x region x product-group grain.",
          "The underlying transaction file has 1,067,371 raw line items. Putting "
          "those in a workbook would produce a several-hundred-megabyte file and "
          "make every SUMIFS crawl, so the line level lives in the Python "
          "pipeline and the workbook carries the aggregate.",
          f"Coverage: {n_months} complete months, four derived geography buckets, "
          "the twelve largest products plus an 'all other products' bucket."]),
        ("Geography buckets are derived, not native",
         ["The raw file has one free-text Country column with 43 distinct values "
          "and no region field of any kind.",
          "United Kingdom / Ireland (EIRE) / Rest of Europe / Rest of World are "
          "buckets defined in analysis/pipeline.py. They are ours, not the data's."]),
        ("Provenance",
         [f"Source: {provenance['dataset_name']} - {provenance['landing_page']}",
          f"File: {provenance['source_url']}",
          f"sha256: {provenance['sha256']}",
          f"Bytes: {provenance['bytes']:,}"]),
    ]
    for head, lines in blocks:
        ws.cell(row=r, column=1, value=head).font = BOLD
        r += 1
        for ln in lines:
            c = ws.cell(row=r, column=1, value=ln)
            c.font = BODY
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws.row_dimensions[r].height = 28
            r += 1
        r += 1

    ws.cell(row=r, column=1, value="Dropped-row ledger").font = BOLD
    r += 1
    _header_row(ws, r, ["Step", "Rows dropped", "Rows remaining", "Why"])
    r += 1
    for s in ledger_steps:
        ws.cell(row=r, column=1, value=s["step"]).font = BODY
        ws.cell(row=r, column=2, value=s["rows_dropped"]).number_format = NUM
        ws.cell(row=r, column=3, value=s["rows_remaining"]).number_format = NUM
        c = ws.cell(row=r, column=4, value=s["why"])
        c.font = BODY
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="Resume claims checked against the data").font = BOLD
    r += 1
    _header_row(ws, r, ["Claim", "Verdict", "What the data says"])
    r += 1
    for cl in claim_notes:
        ws.cell(row=r, column=1, value=cl["resume_claim"]).font = BODY
        ws.cell(row=r, column=2, value=cl["verdict"]).font = BOLD
        c = ws.cell(row=r, column=3, value=cl["real_value"])
        c.font = BODY
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 34
        r += 1

    for col, w in zip("ABCDEF", [46, 16, 62, 16, 16, 16]):
        ws.column_dimensions[col].width = w
